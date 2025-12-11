import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
from openpyxl.styles import PatternFill, Font  # <--- NEW IMPORT FOR STYLING

# --- CONFIGURATION ---
REQUIRED_COLUMNS = [
    "Party Name",
    "Station",
    "Product Name",
    "Qty",
    "Free",
    "Rate",
    "Amount",
    "Tax %"
]

# --- NAME MAPPING (Standard -> List of Variations) ---
PRODUCT_MAPPING = {
    "ALPHALACT-1 400GM": [
        "ALPHALACT -1 4009M",
        "ALPHALACT-I PREM 400GM",
        "Alphalact-1 400gm",
        "ALPHALACT-1 400gm",
        "ALPMALACT-1, 400gm."
    ],
    "ALPHALACT-1 200GM": [
        "ALPHALACT-1200GM",
        "Alphalact-1 200gm"
    ],
    "ALPHALACT-2 400GM": [
        "ALPHALACT 2400GM",
        "ALPHALACT 2 No. 400gm.",
        "ALPHACTT-24009m",
        "ALPHALACT-2 400GM"
    ],
    "ALPHALACT PLUS 400GM": [
        "ALPHALACT PLUS 400GM.",
        "ALPHALACT PLUS 400gm:",
        "Albhalact Plus -1 400gm...",
        "ALPHALACT PRE PLUS 400GM",
        "ALPHALACT PRM PLUS 400GM",
        "ALPHALACT PLUS-1 4009M"
    ],
    "ALPHALACT PLUS 200GM": [
        "ALPHALACT PLUS 2009M",
        "Alphalact-Plus-1200gm",
        "ALPHALACT PRE PLUS 2009M",
        "ALPHALCT PROM PLUS 2009M",
        "ALPHALACT PLUS-1200GM"
    ],
    "ALPHALACT LBW 400GM": [
        "Alphalact LBW 400gm",
        "ALPHALACT LBW 400GM",
        "ALPHALACT PRM LBW 400GM",
        "ALPHALACT LBW 4009M"
    ],
    "ALPHALACT PREMIUM 400GM": [
        "ALPHALACT PREMIUM-400GM",
        "ALPHALACT PREMIUN 4009m"
    ],
    "ALPHALACT PREMIUM 200GM": [
        "ALPHALACT - PREM 2009M.",
        "ALPHALACT PREMIUN 2009m",
        "ALPHALACT PREMIUM 200GM"
    ],
    "ALPHALACT LF 200GM": [
        "ALPHALACT-LF 2009m",
        "ALPHALACT LF-PRE. 2009m.",
        "ALPHALACT-LF 200GM",
        "Alphalact LF 200gm",
        "ALPHALACT LF (2009m 2009m)",
        "ALPHALACT PRMLF 2009M"
    ],
    "ALPHAFIT MOM 200GM": [
        "Alphafit mom 200gm Choc",
        "ALPHAFIT MOm 2009m",
        "ALPHAFIT 2009M Choc."
    ],
    "ALPHAHEALTH PLUS 200GM": [
        "ALPHAHEALTH PLUS 200GM."
    ]
}

# --- FLATTEN MAPPING FOR LOOKUP ---
FLATTENED_MAPPING = {}
for standard_name, variations in PRODUCT_MAPPING.items():
    FLATTENED_MAPPING[standard_name] = standard_name
    FLATTENED_MAPPING[standard_name.upper()] = standard_name
    for var in variations:
        FLATTENED_MAPPING[var] = standard_name
        FLATTENED_MAPPING[var.upper()] = standard_name


# --- HELPER FUNCTIONS ---
def clean_number_str(s):
    if not s: return ""
    return s.replace(',', '').replace('-', '').strip()


def parse_number(s):
    try:
        clean = clean_number_str(s).replace('%', '')
        return float(clean)
    except ValueError:
        return 0.0


def is_numeric_item(s):
    if not s: return False
    clean = clean_number_str(s)
    if clean == '' or len(clean) > 15: return False
    if re.search(r'[a-zA-Z]', clean): return False
    try:
        float(clean)
        return True
    except ValueError:
        return False


def process_pdf(uploaded_file, debug_mode=False):
    all_rows = []
    debug_logs = []
    extracted_title = "Unknown Company"

    EXTRACT_SETTINGS = {
        "x_tolerance": 1,
        "y_tolerance": 5
    }

    current_party = ""
    current_station = ""
    first_line_found = False

    with pdfplumber.open(uploaded_file) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            words = page.extract_words(**EXTRACT_SETTINGS)

            rows = {}
            for word in words:
                y_bucket = round(word['top'] / 5) * 5
                if y_bucket not in rows: rows[y_bucket] = []
                rows[y_bucket].append(word)

            sorted_y = sorted(rows.keys())

            for y in sorted_y:
                line_words = sorted(rows[y], key=lambda w: w['x0'])
                line_text_parts = [w['text'] for w in line_words]
                full_text = " ".join(line_text_parts).strip()

                if not full_text: continue

                # --- TITLE EXTRACTION (First line of Page 1) ---
                if page_num == 1 and not first_line_found:
                    extracted_title = full_text
                    # Clean title
                    extracted_title = re.sub(r'[\\/*?:"<>|]', "", extracted_title)
                    extracted_title = extracted_title.replace("\n", " ").strip()
                    first_line_found = True
                    if debug_mode: debug_logs.append(f"🏢 Company Identified: {extracted_title}")

                if len(full_text) < 3: continue

                if "total" in full_text.lower():
                    if debug_mode: debug_logs.append(f"🚫 Skipped 'Total': {full_text}")
                    continue

                if "Page No" in full_text or "VEDIKA PHARMACY" in full_text or "DESCRIPTION" in full_text:
                    continue

                # --- HEADER DETECTION ---
                is_data_row = False
                numeric_count = sum(1 for w in line_text_parts if is_numeric_item(w))

                if numeric_count >= 3:
                    is_data_row = True

                if not is_data_row:
                    parts = full_text.split("-")
                    if len(parts) >= 2 and not any(is_numeric_item(p) for p in parts):
                        current_station = parts[-1].strip()
                        current_party = "-".join(parts[:-1]).strip()
                    else:
                        current_party = full_text
                        current_station = ""
                    if debug_mode: debug_logs.append(f"🏷️ HEADER: {current_party}")
                    continue

                # --- DATA ROW PARSING ---
                numeric_block = []
                name_parts = []

                for i in range(len(line_text_parts) - 1, -1, -1):
                    word = line_text_parts[i]
                    if word.strip() == '-': continue

                    if is_numeric_item(word):
                        numeric_block.insert(0, parse_number(word))
                    else:
                        name_parts = line_text_parts[:i + 1]
                        break

                qty, rate, amount, tax = 0, 0, 0, 0
                free = ""

                count = len(numeric_block)

                if count >= 5:
                    qty = numeric_block[-5]
                    free = int(numeric_block[-4])
                    rate = numeric_block[-3]
                    amount = numeric_block[-2]
                    tax = numeric_block[-1]
                elif count == 4:
                    qty = numeric_block[-4]
                    free = ""
                    rate = numeric_block[-3]
                    amount = numeric_block[-2]
                    tax = numeric_block[-1]
                elif count == 3:
                    amount = numeric_block[-2]
                    tax = numeric_block[-1]
                    rate = numeric_block[0]
                    free = ""
                else:
                    if debug_mode: debug_logs.append(f"⚠️ SKIPPED (Low Data): {full_text}")
                    continue

                # --- CLEANUP PRODUCT NAME ---
                raw_product_name = " ".join(name_parts).strip()
                raw_product_name = raw_product_name.rstrip(' -')

                match = re.search(r'\s(\d+)\s*-?$', raw_product_name)
                if match:
                    trailing_number = float(match.group(1))
                    if qty == 0 or (trailing_number < 1000 and rate > trailing_number):
                        qty = trailing_number
                        raw_product_name = raw_product_name[:match.start()].strip()
                        raw_product_name = raw_product_name.rstrip(' -')
                        if debug_mode: debug_logs.append(f"🔧 Extracted Qty {qty} from name")

                # --- APPLY MAPPING ---
                final_name = FLATTENED_MAPPING.get(raw_product_name)
                if not final_name:
                    final_name = FLATTENED_MAPPING.get(raw_product_name.upper())
                if not final_name:
                    final_name = raw_product_name

                # Fallback: Extract Party from Name
                if not current_party:
                    if " - " in raw_product_name:
                        split_parts = raw_product_name.split(" - ")
                        current_party = split_parts[0]
                        extracted_name_part = " - ".join(split_parts[1:])
                        mapped_part = FLATTENED_MAPPING.get(extracted_name_part, extracted_name_part)
                        final_name = mapped_part

                all_rows.append({
                    "Party Name": current_party,
                    "Station": current_station,
                    "Product Name": final_name,
                    "Qty": qty,
                    "Free": free,
                    "Rate": rate,
                    "Amount": amount,
                    "Tax %": tax
                })

    return all_rows, debug_logs, extracted_title


def to_excel(df):
    output = io.BytesIO()
    # Use OpenPyXL to allow styling
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Combined Sales Data')

        # Access the workbook and sheet
        workbook = writer.book
        worksheet = writer.sheets['Combined Sales Data']

        # Define Green Style (Light Green Fill + Bold Text)
        # Color code 90EE90 is LightGreen. You can use 00FF00 for bright green.
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        bold_font = Font(bold=True)

        # Iterate over all rows in the sheet
        # Note: OpenPyXL is 1-indexed. Row 1 is the header.
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
            cell = row[0]  # The "Party Name" cell (Column A)

            # Check if this cell contains our "COMPANY:" tag
            if cell.value and str(cell.value).startswith("COMPANY:"):
                # Apply style to the entire row (all columns in this row)
                for cell_in_row in worksheet[cell.row]:
                    cell_in_row.fill = green_fill
                    cell_in_row.font = bold_font

    return output.getvalue()


# --- STREAMLIT UI ---
st.set_page_config(page_title="PDF to Excel Converter", page_icon="📊", layout="wide")

st.title("Customized - PDF to Excel Converter")
st.markdown("Upload **multiple** PDFs. Company Headers will be highlighted in **Green**.")

col1, col2 = st.columns([2, 1])
with col1:
    uploaded_files = st.file_uploader("Choose PDF files", type="pdf", accept_multiple_files=True)
with col2:
    debug_mode = st.checkbox("Show Debug Logs", help="See extracted lines and skipped rows.")

if uploaded_files:
    with st.spinner('Processing files...'):
        all_dfs = []
        all_logs = []

        try:
            for uploaded_file in uploaded_files:
                # 1. Process PDF
                rows, file_logs, file_title = process_pdf(uploaded_file, debug_mode)
                all_logs.extend([f"--- FILE: {uploaded_file.name} ---"] + file_logs)

                if rows:
                    df_file = pd.DataFrame(rows, columns=REQUIRED_COLUMNS)

                    # 2. Cleanup Data (Forward Fill)
                    dash_pattern = r'^[\s\-]+$'
                    df_file["Party Name"] = df_file["Party Name"].replace(dash_pattern, pd.NA, regex=True)
                    df_file["Station"] = df_file["Station"].replace(dash_pattern, pd.NA, regex=True)

                    df_file["Party Name"] = df_file["Party Name"].replace("", pd.NA)
                    df_file["Station"] = df_file["Station"].replace("", pd.NA)

                    df_file["Party Name"] = df_file["Party Name"].ffill()
                    df_file["Station"] = df_file["Station"].ffill()

                    df_file = df_file.fillna("")

                    # 3. Insert Company Name as HEADER ROW
                    header_row_data = {col: "" for col in REQUIRED_COLUMNS}
                    # We prefix with "COMPANY:" so the highlighter can find it later
                    header_row_data["Party Name"] = f"COMPANY: {file_title}"

                    df_header = pd.DataFrame([header_row_data])
                    df_combined_file = pd.concat([df_header, df_file], ignore_index=True)

                    # 4. Add to master list
                    all_dfs.append(df_combined_file)

            # 5. Final Combination
            if all_dfs:
                final_df = pd.concat(all_dfs, ignore_index=True)

                st.success(f"Successfully processed {len(uploaded_files)} files! Total rows: {len(final_df)}")
                st.dataframe(final_df)

                if debug_mode:
                    st.subheader("Debug Logs (All Files)")
                    st.text_area("Log Output", "\n".join(all_logs), height=200)

                excel_data = to_excel(final_df)

                st.download_button(
                    label="📥 Download Combined Excel",
                    data=excel_data,
                    file_name="Combined_Sales_Data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error("No valid data found in the uploaded files.")

        except Exception as e:
            st.error(f"Error: {e}")